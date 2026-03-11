from __future__ import annotations

import argparse
import asyncio
import json
import signal
import subprocess
import time
from pathlib import Path
from typing import Any

import httpx
import yaml
from openpyxl import load_workbook


DEFAULT_QUESTION_SOURCE_PATH = Path("scripts/岱山-指令集-固定问题 -测试问题0311.xlsx")
DEFAULT_HOST = "127.0.0.1"
DEFAULT_PORT = 8000
DEFAULT_USER_ID_PREFIX = "fixed_question_export"
DEFAULT_STARTUP_TIMEOUT = 60
DEFAULT_REQUEST_TIMEOUT = 180.0
DEFAULT_CONFIG_PATH = Path("scripts/fixed_question_test_config.yaml")


class FixedQuestionAnswerExporter:
    """Export answers for fixed questions and archive per-run logs."""

    def __init__(
        self,
        question_source_path: Path,
        host: str = DEFAULT_HOST,
        port: int = DEFAULT_PORT,
        base_url: str | None = None,
        user_id_prefix: str = DEFAULT_USER_ID_PREFIX,
        startup_timeout: int = DEFAULT_STARTUP_TIMEOUT,
        request_timeout: float = DEFAULT_REQUEST_TIMEOUT,
        log_source_path: Path | None = None,
        question_limit: int | None = None,
    ) -> None:
        self.repo_root = Path(__file__).resolve().parents[1]
        self.run_id = time.strftime("%Y%m%d_%H%M%S")
        self.question_source_path = question_source_path
        self.host = host
        self.port = port
        self.user_id_prefix = user_id_prefix
        self.startup_timeout = startup_timeout
        self.request_timeout = request_timeout
        self.base_url = (base_url or f"http://{host}:{port}").rstrip("/")
        self.log_source_path = log_source_path
        self.question_limit = question_limit
        self.server_process: subprocess.Popen[str] | None = None
        self.manage_server = base_url is None
        self.server_log_path = self.question_source_path.with_name(
            f"{self.question_source_path.stem}.{self.run_id}.server.log"
        )
        self.captured_log_path = self.question_source_path.with_name(
            f"{self.question_source_path.stem}.{self.run_id}.captured.log"
        )
        self._log_start_size = 0

    def load_questions(self) -> list[dict[str, Any]]:
        if not self.question_source_path.exists():
            raise FileNotFoundError(f"Question source Excel not found: {self.question_source_path}")

        workbook = load_workbook(self.question_source_path)
        if "Sheet1" not in workbook.sheetnames:
            raise ValueError("Question source Excel must contain Sheet1")

        worksheet = workbook["Sheet1"]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise ValueError("Question source Excel is empty")

        header_map = {str(value).strip(): index + 1 for index, value in enumerate(header_row) if value is not None}
        question_column = header_map.get("问题")
        answer_column = header_map.get("回答")

        if question_column is None:
            raise ValueError("Question source Excel must contain a '问题' column")

        if answer_column is None:
            answer_column = worksheet.max_column + 1
            worksheet.cell(row=1, column=answer_column, value="回答")
            workbook.save(self.question_source_path)

        questions: list[dict[str, Any]] = []
        for row_index in range(2, worksheet.max_row + 1):
            question_value = worksheet.cell(row=row_index, column=question_column).value
            question_text = str(question_value).strip() if question_value is not None else ""
            if not question_text:
                continue
            questions.append(
                {
                    "row_index": row_index,
                    "question": question_text,
                    "answer_column": answer_column,
                }
            )

        workbook.close()

        if not questions:
            raise ValueError("No questions found in Sheet1/问题 column")

        if self.question_limit is not None:
            return questions[: self.question_limit]
        return questions

    def parse_sse_answer(self, lines: list[str]) -> str:
        answer_parts: list[str] = []

        for raw_line in lines:
            line = raw_line.strip()
            if not line.startswith("data: "):
                continue

            payload = line[6:].strip()
            if payload == "[DONE]":
                break

            try:
                event_data = json.loads(payload)
            except json.JSONDecodeError:
                continue

            data_field = event_data.get("data")
            if isinstance(data_field, dict):
                answer_text = str(data_field.get("answer", "") or "")
                if answer_text:
                    answer_parts.append(answer_text)
                    continue

            content_text = str(event_data.get("content", "") or "")
            if content_text:
                answer_parts.append(content_text)

        return "".join(answer_parts).strip()

    def write_results(self, rows: list[dict[str, Any]]) -> Path:
        workbook = load_workbook(self.question_source_path)
        worksheet = workbook["Sheet1"]

        for row in rows:
            worksheet.cell(
                row=int(row["row_index"]),
                column=int(row["answer_column"]),
                value=row.get("answer", ""),
            )

        workbook.save(self.question_source_path)
        workbook.close()
        return self.question_source_path

    def mark_log_start(self) -> None:
        if self.log_source_path is None:
            return
        if not self.log_source_path.exists():
            self._log_start_size = 0
            return
        self._log_start_size = self.log_source_path.stat().st_size

    def capture_run_logs(self) -> Path | None:
        if self.log_source_path is None:
            return None

        self.captured_log_path.parent.mkdir(parents=True, exist_ok=True)
        if not self.log_source_path.exists():
            self.captured_log_path.write_text(
                f"Log source not found: {self.log_source_path}\n",
                encoding="utf-8",
            )
            return self.captured_log_path

        current_size = self.log_source_path.stat().st_size
        start_size = self._log_start_size
        if current_size < start_size:
            start_size = 0

        with self.log_source_path.open("r", encoding="utf-8", errors="ignore") as source:
            source.seek(start_size)
            content = source.read()

        self.captured_log_path.write_text(content, encoding="utf-8")
        return self.captured_log_path

    def start_server(self) -> None:
        if not self.manage_server:
            return
        if self.server_process is not None:
            return

        self.question_source_path.parent.mkdir(parents=True, exist_ok=True)
        log_file = self.server_log_path.open("w", encoding="utf-8")

        command = [
            "uv",
            "run",
            "--project",
            str(self.repo_root / "src" / "rag_stream"),
            "python",
            "-m",
            "uvicorn",
            "main:app",
            "--host",
            self.host,
            "--port",
            str(self.port),
            "--app-dir",
            str(self.repo_root / "src" / "rag_stream"),
        ]

        self.server_process = subprocess.Popen(
            command,
            stdout=log_file,
            stderr=subprocess.STDOUT,
            text=True,
            cwd=self.repo_root,
        )

    async def wait_for_server(self) -> None:
        deadline = time.time() + self.startup_timeout
        last_error = ""

        async with httpx.AsyncClient(timeout=2.0) as client:
            while time.time() < deadline:
                try:
                    response = await client.get(f"{self.base_url}/health")
                    if response.status_code == 200:
                        return
                    last_error = f"Unexpected health status: {response.status_code}"
                except Exception as error:
                    last_error = str(error)

                if self.server_process is not None and self.server_process.poll() is not None:
                    raise RuntimeError(
                        f"Server exited before becoming healthy. Check log: {self.server_log_path}"
                    )

                await asyncio.sleep(1)

        raise RuntimeError(
            f"Server did not become healthy within {self.startup_timeout}s: {last_error}"
        )

    def stop_server(self) -> None:
        if self.server_process is None:
            return

        try:
            self.server_process.send_signal(signal.SIGTERM)
            self.server_process.wait(timeout=10)
        except subprocess.TimeoutExpired:
            self.server_process.kill()
            self.server_process.wait(timeout=5)
        finally:
            self.server_process = None

    async def fetch_answer(
        self,
        client: httpx.AsyncClient,
        question: str,
        question_index: int,
    ) -> str:
        payload = {
            "question": question,
            "user_id": f"{self.user_id_prefix}_{self.run_id}_{question_index:04d}",
            "stream": True,
        }
        sse_lines: list[str] = []

        async with client.stream(
            "POST",
            f"{self.base_url}/api/general",
            json=payload,
            timeout=self.request_timeout,
        ) as response:
            if response.status_code != 200:
                error_text = await response.aread()
                raise RuntimeError(
                    f"HTTP {response.status_code}: {error_text.decode('utf-8', errors='ignore')}"
                )

            async for line in response.aiter_lines():
                if line:
                    sse_lines.append(line)

        answer = self.parse_sse_answer(sse_lines)
        if answer:
            return answer
        raise RuntimeError("Received empty answer stream")

    async def export(self) -> Path:
        questions = self.load_questions()
        self.question_source_path.parent.mkdir(parents=True, exist_ok=True)
        self.mark_log_start()
        self.start_server()
        await self.wait_for_server()

        rows: list[dict[str, Any]] = []
        try:
            async with httpx.AsyncClient() as client:
                for index, question_item in enumerate(questions, start=1):
                    question = str(question_item["question"])
                    print(f"[{index}/{len(questions)}] Asking: {question}")
                    try:
                        answer = await self.fetch_answer(client, question, index)
                    except Exception as error:
                        answer = f"[ERROR] {error}"
                    rows.append({**question_item, "answer": answer})

            output_path = self.write_results(rows)
            captured_log_path = self.capture_run_logs()
            print(f"Updated {len(rows)} answers in: {output_path}")
            if self.manage_server:
                print(f"Server log saved to: {self.server_log_path}")
            if captured_log_path is not None:
                print(f"Captured app log saved to: {captured_log_path}")
            return output_path
        finally:
            self.stop_server()


def load_script_config(config_path: Path) -> dict[str, Any]:
    if not config_path.exists():
        raise FileNotFoundError(f"Config file not found: {config_path}")

    with config_path.open("r", encoding="utf-8") as file:
        config = yaml.safe_load(file) or {}

    if not isinstance(config, dict):
        raise ValueError("Config file must be a YAML object")

    return config


def _normalize_optional_path(value: Any) -> Path | None:
    if value in (None, ""):
        return None
    return Path(str(value))


def build_exporter_from_config(config: dict[str, Any]) -> FixedQuestionAnswerExporter:
    return FixedQuestionAnswerExporter(
        question_source_path=Path(str(config.get("question_source_path", DEFAULT_QUESTION_SOURCE_PATH))),
        host=str(config.get("host", DEFAULT_HOST)),
        port=int(config.get("port", DEFAULT_PORT)),
        base_url=str(config["base_url"]).rstrip("/") if config.get("base_url") else None,
        user_id_prefix=str(config.get("user_id_prefix", DEFAULT_USER_ID_PREFIX)),
        startup_timeout=int(config.get("startup_timeout", DEFAULT_STARTUP_TIMEOUT)),
        request_timeout=float(config.get("request_timeout", DEFAULT_REQUEST_TIMEOUT)),
        log_source_path=_normalize_optional_path(config.get("log_source_path")),
        question_limit=int(config["question_limit"]) if config.get("question_limit") not in (None, "") else None,
    )


async def _async_main() -> int:
    parser = argparse.ArgumentParser(
        description="Read fixed questions from Excel, request /api/general, and overwrite the Excel 回答 column."
    )
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG_PATH,
        help="YAML config path. Defaults to scripts/fixed_question_test_config.yaml.",
    )
    parser.add_argument(
        "--question-limit",
        type=int,
        default=None,
        help="Optional one-off override for the number of Excel questions to run.",
    )
    args = parser.parse_args()

    config = load_script_config(args.config)
    exporter = build_exporter_from_config(config)

    if args.question_limit is not None:
        exporter.question_limit = args.question_limit

    await exporter.export()
    return 0


def main() -> int:
    return asyncio.run(_async_main())


if __name__ == "__main__":
    raise SystemExit(main())
